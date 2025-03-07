from collections import deque
import cv2
import numpy as np
from scipy.optimize import linear_sum_assignment
from ultralytics import YOLO
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import io
import logging

# Configure logging for debugging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class WasteTracker:
    def __init__(self, trash_model_path, video_path, evidence_path, excel_output_path, 
                 distance_threshold=150, max_inactive_frames=90, temporal_window=10,
                 min_holding_frames=15, min_disposal_frames=20, min_throw_frames=5,  # New parameter for throwing
                 camera_location="Camera_01", display_width=800, display_height=600):  # New parameters for display size
        # Initialize models and parameters
        try:
            self.yolo_model = YOLO("yolov8m.pt")  # For humans and vehicles
            self.trash_model = YOLO(trash_model_path)  # For trash
        except Exception as e:
            raise ValueError(f"Failed to load YOLO models: {e}")

        self.video_path = video_path
        self.evidence_path = evidence_path
        self.excel_output_path = excel_output_path
        self.distance_threshold = distance_threshold
        self.max_inactive_frames = max_inactive_frames
        self.temporal_window = temporal_window
        self.min_holding_frames = min_holding_frames
        self.min_disposal_frames = min_disposal_frames
        self.min_throw_frames = min_throw_frames  # For brief trash appearances during throwing
        self.camera_location = camera_location
        
        self.tracking_data = {}
        self.next_id = 1
        self.frame_count = 0
        self.events_data = []

        # New parameters for controlling display size
        self.display_width = display_width
        self.display_height = display_height

        # Create directories if they don't exist
        os.makedirs(self.evidence_path, exist_ok=True)
        os.makedirs(self.excel_output_path, exist_ok=True)

    def calculate_center(self, bbox):
        x1, y1, x2, y2 = bbox
        return int((x1 + x2) / 2), int((y1 + y2) / 2)

    def compute_distance(self, center1, center2):
        return np.sqrt((center1[0] - center2[0])**2 + (center1[1] - center2[1])**2)

    def calculate_velocity(self, coordinates, bbox, frame_interval=2):
        """Calculate vehicle velocity, normalizing for perspective using bounding box area."""
        if len(coordinates) < frame_interval + 1:
            return 0, bbox[2] * bbox[3]  # Return 0 velocity and bbox area if not enough data
        prev_center = coordinates[-frame_interval - 1]
        curr_center = coordinates[-1]
        distance = self.compute_distance(prev_center, curr_center)
        pixel_velocity = distance / frame_interval  # Pixels per frame
        bbox_area = bbox[2] * bbox[3]  # Area = width * height
        # Normalize velocity using two methods for testing: option 1 (bbox_area ** 0.25), option 2 (capped at 10000 ** 0.5)
        normalized_velocity1 = pixel_velocity / (bbox_area ** 0.25 if bbox_area > 0 else 1)  # Reduced impact of large areas
        normalized_velocity2 = pixel_velocity / (min(bbox_area, 10000) ** 0.5 if bbox_area > 0 else 1)  # Capped normalization
        # Use the higher of the two for better accuracy (you can test and choose one later)
        normalized_velocity = max(normalized_velocity1, normalized_velocity2)
        return normalized_velocity, bbox_area

    def initialize_track(self, det, current_frame):
        track_id = self.next_id
        self.tracking_data[track_id] = {
            'class_id': det['class_id'],
            'bbox': det['bbox'],
            'coordinates': deque([self.calculate_center(det['bbox'])], maxlen=20),  # Larger buffer for velocity
            'last_seen': current_frame,
            'proximity_buffer': deque([0] * self.temporal_window, maxlen=self.temporal_window),
            'throw_buffer': deque([0] * self.min_throw_frames, maxlen=self.min_throw_frames),  # New buffer for throwing
            'frame_buffer': deque(maxlen=10),
            'state': 'IDLE',
            'no_trash_count': 0,
            'type': 'human' if det['class_id'] == 0 else 'trash' if det['class_id'] == 1 else 'vehicle',
            'velocity_history': deque([0] * 5, maxlen=5),  # Track normalized velocity
            'bbox_area_history': deque([det['bbox'][2] * det['bbox'][3]] * 5, maxlen=5)  # Track bounding box area
        }
        det['id'] = track_id
        self.next_id += 1
        logger.debug(f"Initialized track {track_id} as {self.tracking_data[track_id]['type']}")

    def update_track(self, track_id, det, current_frame):
        self.tracking_data[track_id]['bbox'] = det['bbox']
        center = self.calculate_center(det['bbox'])
        self.tracking_data[track_id]['coordinates'].append(center)
        self.tracking_data[track_id]['last_seen'] = current_frame
        if self.tracking_data[track_id]['type'] == 'vehicle':
            # Calculate normalized velocity and bounding box area
            normalized_velocity, bbox_area = self.calculate_velocity(
                self.tracking_data[track_id]['coordinates'], det['bbox']
            )
            self.tracking_data[track_id]['velocity_history'].append(normalized_velocity)
            self.tracking_data[track_id]['bbox_area_history'].append(bbox_area)

    def compute_distance_matrix(self, detections, tracking_data):
        if not detections or not tracking_data:
            return np.array([])
        distance_matrix = []
        for det in detections:
            det_center = self.calculate_center(det['bbox'])
            distances = [self.compute_distance(det_center, track['coordinates'][-1]) 
                         for track in tracking_data.values()]
            distance_matrix.append(distances)
        return np.array(distance_matrix)

    def assign_ids(self, detections, current_frame):
        if not detections:
            return

        if not self.tracking_data:
            for det in detections:
                self.initialize_track(det, current_frame)
            return

        distance_matrix = self.compute_distance_matrix(detections, self.tracking_data)
        if distance_matrix.size == 0:
            for det in detections:
                self.initialize_track(det, current_frame)
            return

        row_indices, col_indices = linear_sum_assignment(distance_matrix)
        assigned_ids = set()

        for row, col in zip(row_indices, col_indices):
            if col >= len(self.tracking_data) or row >= len(detections):
                continue
            if distance_matrix[row, col] < self.distance_threshold:
                det = detections[row]
                track_id = list(self.tracking_data.keys())[col]
                if track_id not in assigned_ids:
                    det['id'] = track_id
                    self.update_track(track_id, det, current_frame)
                    assigned_ids.add(track_id)

        for det in detections:
            if 'id' not in det:
                self.initialize_track(det, current_frame)

        self.clean_inactive_tracks(current_frame)

    def clean_inactive_tracks(self, current_frame):
        inactive_tracks = [tid for tid, data in self.tracking_data.items()
                          if current_frame - data['last_seen'] > self.max_inactive_frames]
        for tid in inactive_tracks:
            del self.tracking_data[tid]
            logger.debug(f"Cleaned inactive track {tid}")

    def detect_objects(self, frame):
        """Detect humans, vehicles, and trash with full accuracy."""
        detections = []
        try:
            # YOLOv8 for humans and vehicles (class 0 for person, 2/7 for vehicles)
            yolo_results = self.yolo_model(frame, conf=0.5, iou=0.4)
            for result in yolo_results:
                for box in result.boxes:
                    bbox = box.xyxy[0].cpu().numpy()
                    class_id = int(box.cls[0].cpu().numpy())
                    if class_id in [0, 2, 7]:  # Person or vehicle
                        detections.append({'bbox': bbox, 'class_id': class_id})

            # Custom model for trash detection (class 1 for trash)
            trash_results = self.trash_model(frame, conf=0.5, iou=0.4)
            for result in trash_results:
                for box in result.boxes:
                    bbox = box.xyxy[0].cpu().numpy()
                    class_id = int(box.cls[0].cpu().numpy())
                    if class_id == 1:  # Trash
                        detections.append({'bbox': bbox, 'class_id': 1})
        except Exception as e:
            logger.error(f"Detection error: {e}")
        return detections

    def process_proximity_logic(self, detections, frame):
        for track_id, track_data in self.tracking_data.items():
            if track_data['type'] not in ['human', 'vehicle']:
                continue

            entity_center = track_data['coordinates'][-1]
            is_trash_near = False

            for det in detections:
                if det['class_id'] == 1:  # Trash
                    trash_center = self.calculate_center(det['bbox'])
                    if self.compute_distance(entity_center, trash_center) < self.distance_threshold:
                        is_trash_near = True
                        break

            track_data['proximity_buffer'].append(1 if is_trash_near else 0)
            track_data['throw_buffer'].append(1 if is_trash_near else 0)  # Update throw buffer for vehicles
            if self.frame_count % 2 == 0:  # Maintain every 2nd frame for full accuracy
                track_data['frame_buffer'].append(frame.copy())

            consecutive_trash_frames = sum(track_data['proximity_buffer'])
            consecutive_throw_frames = sum(track_data['throw_buffer'])
            entity_type = track_data['type']

            if track_data['type'] == 'vehicle':
                # Calculate average normalized velocity and bounding box area for distance-based analysis
                velocities = list(track_data['velocity_history'])
                bbox_areas = list(track_data['bbox_area_history'])
                avg_velocity = np.mean(velocities) if velocities else 0
                avg_bbox_area = np.mean(bbox_areas) if bbox_areas else 0

                # Define distance bands based on bounding box area (adjusted thresholds for your footage)
                if avg_bbox_area > 100000:  # Near (large vehicles close to camera, e.g., 100,000–200,000 pixels)
                    is_stopped = avg_velocity < 0.2  # Lower threshold for near vehicles
                    is_moving = avg_velocity > 1.0    # Higher threshold for near vehicles
                    is_slowed = 0.2 <= avg_velocity < 1.0  # Slowed range for near vehicles
                elif 20000 < avg_bbox_area <= 100000:  # Mid (medium-sized vehicles, e.g., 20,000–50,000 pixels)
                    is_stopped = avg_velocity < 0.4
                    is_moving = avg_velocity > 2.0
                    is_slowed = 0.4 <= avg_velocity < 2.0  # Slowed range for mid vehicles
                else:  # Far (small, distant vehicles, e.g., 5,000–10,000 pixels)
                    is_stopped = avg_velocity < 0.8
                    is_moving = avg_velocity > 4.0
                    is_slowed = 0.8 <= avg_velocity < 4.0  # Slowed range for far vehicles

                if track_data['state'] == 'IDLE':
                    if is_stopped and consecutive_trash_frames >= self.min_holding_frames:
                        track_data['state'] = 'STOPPED_UNLOADING'
                        track_data['no_trash_count'] = 0
                        logger.info(f"Vehicle ID {track_id} confirmed to be stopped and unloading trash (near/mid/far: {avg_bbox_area})")
                    elif is_slowed and consecutive_throw_frames >= self.min_throw_frames:
                        track_data['state'] = 'SLOWING_THROW'
                        track_data['no_trash_count'] = 0
                        logger.info(f"Vehicle ID {track_id} detected as slowing down to throw waste (near/mid/far: {avg_bbox_area})")

                elif track_data['state'] == 'STOPPED_UNLOADING':
                    if is_trash_near:
                        track_data['no_trash_count'] = 0
                    else:
                        track_data['no_trash_count'] += 1
                        if track_data['no_trash_count'] >= self.min_disposal_frames and is_moving:
                            track_data['state'] = 'TRASH_DISPOSED'
                            logger.info(f"Vehicle ID {track_id} disposed trash and left (near/mid/far: {avg_bbox_area})")
                            self.save_evidence(track_id, track_data['frame_buffer'], frame, entity_type, 'STOPPED_UNLOADING')
                            track_data['frame_buffer'].clear()
                            track_data['proximity_buffer'].clear()
                            track_data['throw_buffer'].clear()
                            track_data['no_trash_count'] = 0

                elif track_data['state'] == 'SLOWING_THROW':
                    if is_trash_near:
                        track_data['no_trash_count'] = 0
                    else:
                        track_data['no_trash_count'] += 1
                        if track_data['no_trash_count'] >= self.min_disposal_frames and is_moving:
                            track_data['state'] = 'TRASH_DISPOSED'
                            logger.info(f"Vehicle ID {track_id} threw waste and resumed moving (near/mid/far: {avg_bbox_area})")
                            self.save_evidence(track_id, track_data['frame_buffer'], frame, entity_type, 'SLOWING_THROW')
                            track_data['frame_buffer'].clear()
                            track_data['proximity_buffer'].clear()
                            track_data['throw_buffer'].clear()
                            track_data['no_trash_count'] = 0

            elif track_data['type'] == 'human':
                if track_data['state'] == 'IDLE':
                    if consecutive_trash_frames >= self.min_holding_frames:
                        track_data['state'] = 'HOLDING_TRASH'
                        track_data['no_trash_count'] = 0
                        logger.info(f"Human ID {track_id} confirmed to be holding trash")

                elif track_data['state'] == 'HOLDING_TRASH':
                    if is_trash_near:
                        track_data['no_trash_count'] = 0
                    else:
                        track_data['no_trash_count'] += 1
                        if track_data['no_trash_count'] >= self.min_disposal_frames:
                            track_data['state'] = 'TRASH_DISPOSED'
                            logger.info(f"Human ID {track_id} disposed trash")
                            self.save_evidence(track_id, track_data['frame_buffer'], frame, entity_type)
                            track_data['frame_buffer'].clear()
                            track_data['proximity_buffer'].clear()
                            track_data['no_trash_count'] = 0

    def save_evidence(self, entity_id, frame_buffer, current_frame, entity_type, disposal_type=None):
        """Save evidence with initial and final frames, including disposal type for vehicles."""
        folder_name = f"{entity_type}_{entity_id}"
        entity_folder = os.path.join(self.evidence_path, folder_name)
        os.makedirs(entity_folder, exist_ok=True)

        first_frame = frame_buffer[0].copy() if frame_buffer else current_frame.copy()
        final_frame = current_frame.copy()

        # Draw bounding boxes on evidence frames
        x1, y1, x2, y2 = map(int, self.tracking_data[entity_id]['bbox'])
        color = (0, 0, 255) if entity_type == 'human' else (255, 0, 0)  # Red for humans, Blue for vehicles
        for frame in [first_frame, final_frame]:
            cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
            label = f"{entity_type.capitalize()} ID {entity_id}"
            if disposal_type:
                label += f" - {disposal_type}"
            cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

        # Convert to PIL Images for Excel
        first_frame_pil = Image.fromarray(cv2.cvtColor(first_frame, cv2.COLOR_BGR2RGB))
        final_frame_pil = Image.fromarray(cv2.cvtColor(final_frame, cv2.COLOR_BGR2RGB))
        max_size = (800, 600)

        def resize_image(img, max_size):
            ratio = min(max_size[0]/img.size[0], max_size[1]/img.size[1])
            new_size = tuple(int(dim * ratio) for dim in img.size)
            return img.resize(new_size, Image.Resampling.LANCZOS if hasattr(Image, 'Resampling') else Image.ANTIALIAS)

        first_frame_pil = resize_image(first_frame_pil, max_size)
        final_frame_pil = resize_image(final_frame_pil, max_size)

        # Store event data
        current_time = datetime.now()
        event_data = {
            'Date': current_time.strftime('%Y-%m-%d'),
            'Time': current_time.strftime('%H:%M:%S'),
            'Location': self.camera_location,
            'Entity_Type': entity_type,
            'Entity_ID': entity_id,
            'Evidence_Folder': entity_folder,
            'Initial_Frame': first_frame_pil,
            'Final_Frame': final_frame_pil
        }
        if disposal_type:
            event_data['Disposal_Type'] = disposal_type
        self.events_data.append(event_data)
        logger.info(f"Saved evidence for {entity_type.capitalize()} ID {entity_id} - {disposal_type if disposal_type else 'TRASH_DISPOSED'}")

    def export_events(self):
        """Export events with images and disposal type to Excel."""
        if not self.events_data:
            logger.warning("No events to export")
            return None

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'waste_disposal_events_{timestamp}.xlsx'
        full_path = os.path.join(self.excel_output_path, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Disposal Events"

        headers = ['Date', 'Time', 'Location', 'Entity Type', 'Entity ID', 'Disposal Type', 'Evidence Folder', 'Initial Frame', 'Final Frame']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 40

        for row, event in enumerate(self.events_data, 2):
            ws.cell(row=row, column=1, value=event['Date'])
            ws.cell(row=row, column=2, value=event['Time'])
            ws.cell(row=row, column=3, value=event['Location'])
            ws.cell(row=row, column=4, value=event['Entity_Type'])
            ws.cell(row=row, column=5, value=event['Entity_ID'])
            ws.cell(row=row, column=6, value=event.get('Disposal_Type', 'TRASH_DISPOSED'))
            ws.cell(row=row, column=7, value=event['Evidence_Folder'])

            ws.row_dimensions[row].height = 180
            for col, img_key in enumerate(['Initial_Frame', 'Final_Frame'], 8):
                img_buffer = io.BytesIO()
                event[img_key].save(img_buffer, format='PNG')
                img = ExcelImage(img_buffer)
                ws.add_image(img, f'{chr(64+col)}{row}')

        try:
            wb.save(full_path)
            logger.info(f"Events exported to: {full_path}")
            return full_path
        except Exception as e:
            logger.error(f"Failed to export events: {e}")
            return None

    def visualize_frame(self, frame, detections):
        """Visualize with distinct colors, state messages, and vehicle velocity for vehicles."""
        # Resize the frame for display (maintain aspect ratio)
        aspect_ratio = frame.shape[1] / frame.shape[0]
        new_width = self.display_width
        new_height = int(self.display_width / aspect_ratio) if aspect_ratio > 1 else int(self.display_height * aspect_ratio)
        if new_height > self.display_height:
            new_height = self.display_height
            new_width = int(self.display_height * aspect_ratio)
        resized_frame = cv2.resize(frame, (new_width, new_height))

        for det in detections:
            x1, y1, x2, y2 = map(int, det['bbox'])
            # Scale bounding box coordinates for the resized frame
            scale_x = new_width / frame.shape[1]
            scale_y = new_height / frame.shape[0]
            scaled_x1, scaled_y1 = int(x1 * scale_x), int(y1 * scale_y)
            scaled_x2, scaled_y2 = int(x2 * scale_x), int(y2 * scale_y)

            if det['class_id'] == 1:  # Trash
                color = (0, 255, 0)  # Green
                label = f"ID {det.get('id', 'N/A')} - Trash"
            elif det['class_id'] == 0:  # Human
                color = (0, 0, 255)  # Red
                label = f"ID {det.get('id', 'N/A')} - Human"
            else:  # Vehicle (class 2 or 7)
                color = (255, 0, 0)  # Blue
                track_id = det.get('id', 'N/A')
                if track_id in self.tracking_data and self.tracking_data[track_id]['type'] == 'vehicle':
                    velocities = list(self.tracking_data[track_id]['velocity_history'])
                    bbox_areas = list(self.tracking_data[track_id]['bbox_area_history'])
                    avg_velocity = np.mean(velocities) if velocities else 0
                    avg_bbox_area = np.mean(bbox_areas) if bbox_areas else 0
                    label = f"ID {track_id} - Vehicle (Vel: {avg_velocity:.1f}, Area: {int(avg_bbox_area)})"
                else:
                    label = f"ID {track_id} - Vehicle"

            cv2.rectangle(resized_frame, (scaled_x1, scaled_y1), (scaled_x2, scaled_y2), color, 2)
            cv2.putText(resized_frame, label, (scaled_x1, scaled_y1 - 40), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)  # Adjusted for longer label

        # Display state messages for humans and vehicles on the resized frame
        y_offset = 30
        for track_id, track_data in self.tracking_data.items():
            if track_data['type'] in ['human', 'vehicle']:
                state_text = f"{track_data['type'].capitalize()} ID {track_id}: {track_data['state']}"
                cv2.putText(resized_frame, state_text, (10, y_offset), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
                y_offset += 30

        return resized_frame

    def process_video(self, video_path=None):
        """Process video with full accuracy, resizing the display for better visibility."""
        if video_path is None:
            video_path = self.video_path  # Use instance variable if not provided
        try:
            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():
                raise ValueError(f"Failed to open video: {video_path}")

            current_frame = 0
            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    break

                # Process every 2nd frame for full accuracy (no skip increase)
                if current_frame % 2 == 0:
                    detections = self.detect_objects(frame)
                    self.assign_ids(detections, current_frame)
                    self.process_proximity_logic(detections, frame)
                    resized_frame = self.visualize_frame(frame, detections)

                    cv2.imshow("Tracking", resized_frame)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

                current_frame += 1
                self.frame_count += 1

        except Exception as e:
            logger.error(f"Video processing error: {e}")
        finally:
            cap.release()
            cv2.destroyAllWindows()
            result = self.export_events()
            logger.info(f"Processing complete. Excel file: {result}")
            return result

if __name__ == "__main__":
    # File paths
    trash_model_path = r"C:\Users\DELL\Desktop\trash\veh_pavi\vehidet\100epochv2.pt"
    video_path = r"C:\Users\DELL\Desktop\trash\veh_pavi\vehidet\mech1.mp4"
    evidence_path = r"C:\Users\DELL\Desktop\trash\veh_pavi\vehidet\evidence"
    excel_output_path = r"C:\Users\DELL\Desktop\trash\veh_pavi\vehidet\reports"
    
    # Initialize tracker with customizable display size (adjust these values as needed)
    tracker = WasteTracker(
        trash_model_path=trash_model_path,
        video_path=video_path,
        evidence_path=evidence_path,
        excel_output_path=excel_output_path,
        distance_threshold=150,
        max_inactive_frames=90,
        temporal_window=10,
        min_holding_frames=15,  # Adjusted to match your reference
        min_disposal_frames=20,  # Adjusted to match your reference
        min_throw_frames=5,     # New parameter for throwing detection
        camera_location="Camera_01",
        display_width=1280,  # Adjust this to make the video less zoomed-in (e.g., 800, 640, etc.)
        display_height=720   # Adjust this to match your desired aspect ratio
    )
    
    # Process video
    tracker.process_video()