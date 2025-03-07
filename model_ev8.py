from collections import deque
import cv2
import numpy as np
from scipy.optimize import linear_sum_assignment
from ultralytics import YOLO
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from PIL import Image
import io

class WasteTracker:
    def __init__(self, trash_model_path, evidence_path, distance_threshold, pair_consistency_threshold, max_inactive_frames, 
                 temporal_window=10, temporal_threshold=7, min_holding_frames=15, min_disposal_frames=20, camera_location=""):
        # Initialize models and basic parameters
        self.yolo_model = YOLO("yolov8m.pt")
        self.trash_model = YOLO(trash_model_path)
        
        self.evidence_path = evidence_path
        self.distance_threshold = distance_threshold
        self.pair_consistency_threshold = pair_consistency_threshold
        self.max_inactive_frames = max_inactive_frames
        self.temporal_window = temporal_window
        self.temporal_threshold = temporal_threshold
        self.min_holding_frames = min_holding_frames
        self.min_disposal_frames = min_disposal_frames
        self.tracking_data = {}
        self.next_id = 1
        self.frame_count = 0
        self.camera_location = camera_location
        
        # Initialize event storage with image data
        self.events_data = []
        
        os.makedirs(self.evidence_path, exist_ok=True)

    def calculate_center(self, bbox):
        x1, y1, x2, y2 = bbox
        center_x = (x1 + x2) / 2
        center_y = (y1 + y2) / 2
        return int(center_x), int(center_y)

    def compute_distance(self, center1, center2):
        return np.sqrt((center1[0] - center2[0])**2 + (center1[1] - center2[1])**2)

    def assign_ids(self, detections, current_frame):
        if not self.tracking_data:
            for det in detections:
                self.initialize_track(det, current_frame)
            return

        if not detections:
            return

        distance_matrix = self.compute_distance_matrix(detections, self.tracking_data)

        if distance_matrix.size == 0:
            for det in detections:
                self.initialize_track(det, current_frame)
            return

        row_indices, col_indices = linear_sum_assignment(distance_matrix)

        assigned_ids = set()
        for row, col in zip(row_indices, col_indices):
            if col >= len(self.tracking_data) or row >= len(detections):
                continue

            if distance_matrix[row, col] < self.distance_threshold:
                det = detections[row]
                track_id = list(self.tracking_data.keys())[col]
                if track_id not in assigned_ids:
                    det['id'] = track_id
                    self.update_track(track_id, det, current_frame)
                    assigned_ids.add(track_id)

        for det in detections:
            if 'id' not in det:
                self.initialize_track(det, current_frame)

        self.clean_inactive_tracks(current_frame)

    def initialize_track(self, det, current_frame):
        track_id = self.next_id
        self.tracking_data[track_id] = {
            'class_id': det['class_id'],
            'bbox': det['bbox'],
            'coordinates': deque([self.calculate_center(det['bbox'])], maxlen=10),
            'last_seen': current_frame,
            'proximity_buffer': deque([0] * self.temporal_window, maxlen=self.temporal_window),
            'frame_buffer': deque(maxlen=10),
            'state': 'IDLE',
            'no_trash_count': 0,
            'type': 'human' if det['class_id'] == 0 else 'vehicle' if det['class_id'] == 2 else 'trash'
        }
        det['id'] = track_id
        self.next_id += 1

    def update_track(self, track_id, det, current_frame):
        self.tracking_data[track_id]['bbox'] = det['bbox']
        self.tracking_data[track_id]['coordinates'].append(self.calculate_center(det['bbox']))
        self.tracking_data[track_id]['last_seen'] = current_frame

    def compute_distance_matrix(self, detections, tracking_data):
        distance_matrix = []
        for det in detections:
            det_center = self.calculate_center(det['bbox'])
            distances = []
            for track in tracking_data.values():
                track_center = track['coordinates'][-1]
                distance = self.compute_distance(det_center, track_center)
                distances.append(distance)
            distance_matrix.append(distances)
        return np.array(distance_matrix)

    def record_disposal_event(self, entity_id, entity_type, evidence_folder):
        """Record a disposal event in the events DataFrame"""
        current_time = datetime.now()
        
        # Create event record
        event_data = {
            'Date': current_time.strftime('%Y-%m-%d'),
            'Time': current_time.strftime('%H:%M:%S'),
            'Location': self.camera_location,
            'Entity_Type': entity_type,
            'Entity_ID': entity_id,
            'Evidence_Folder': evidence_folder,
            'Initial_Frame': os.path.join(evidence_folder, 'frame_000.jpg'),
            'Final_Frame': os.path.join(evidence_folder, 'final_frame.jpg')
        }
        
        # Append to DataFrame
        self.events_df = pd.concat([self.events_df, pd.DataFrame([event_data])], ignore_index=True)

    def save_evidence(self, entity_id, frame_buffer, current_frame, entity_type):
        """Save evidence and store images for Excel export"""
        folder_name = f"vehicle_{entity_id}" if entity_type == 'vehicle' else f"human_{entity_id}"
        entity_folder = os.path.join(self.evidence_path, folder_name)
        os.makedirs(entity_folder, exist_ok=True)
        
        # Save first and final frames for the event
        first_frame = frame_buffer[0].copy()
        final_frame = current_frame.copy()
        
        # Draw bounding boxes
        x1, y1, x2, y2 = map(int, self.tracking_data[entity_id]['bbox'])
        for frame in [first_frame, final_frame]:
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 2)
            cv2.putText(frame, f"{entity_type.capitalize()} ID {entity_id}", 
                       (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
        
        # Convert frames to PIL Images
        first_frame_pil = Image.fromarray(cv2.cvtColor(first_frame, cv2.COLOR_BGR2RGB))
        final_frame_pil = Image.fromarray(cv2.cvtColor(final_frame, cv2.COLOR_BGR2RGB))
        
        # Resize images for Excel (to prevent large file sizes)
        max_size = (800, 600)
        
        # Version-compatible image resizing
        def resize_image(img, max_size):
            ratio = min(max_size[0]/img.size[0], max_size[1]/img.size[1])
            new_size = tuple(int(dim * ratio) for dim in img.size)
            try:
                # Try newer Pillow version syntax
                return img.resize(new_size, Image.Resampling.LANCZOS)
            except AttributeError:
                # Fall back to older Pillow version syntax
                return img.resize(new_size, Image.ANTIALIAS)
        
        first_frame_pil = resize_image(first_frame_pil, max_size)
        final_frame_pil = resize_image(final_frame_pil, max_size)
        
        # Store event data
        current_time = datetime.now()
        event_data = {
            'Date': current_time.strftime('%Y-%m-%d'),
            'Time': current_time.strftime('%H:%M:%S'),
            'Location': self.camera_location,
            'Entity_Type': entity_type,
            'Entity_ID': entity_id,
            'Initial_Frame': first_frame_pil,
            'Final_Frame': final_frame_pil
        }
        
        self.events_data.append(event_data)
        print(f"Saved evidence for {entity_type.capitalize()} ID {entity_id}")

    def export_events(self, output_path):
        """Export events with embedded images to Excel"""
        if not self.events_data:
            print("No events to export")
            return None
            
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'waste_disposal_events_{timestamp}.xlsx'
        full_path = os.path.join(output_path, filename)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Disposal Events"
        
        # Set column headers
        headers = ['Date', 'Time', 'Location', 'Entity Type', 'Entity ID', 'Initial Frame', 'Final Frame']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Date
        ws.column_dimensions['B'].width = 15  # Time
        ws.column_dimensions['C'].width = 20  # Location
        ws.column_dimensions['D'].width = 15  # Entity Type
        ws.column_dimensions['E'].width = 15  # Entity ID
        ws.column_dimensions['F'].width = 40  # Initial Frame
        ws.column_dimensions['G'].width = 40  # Final Frame
        
        # Add data and images
        for row, event in enumerate(self.events_data, 2):
            # Add text data
            ws.cell(row=row, column=1, value=event['Date'])
            ws.cell(row=row, column=2, value=event['Time'])
            ws.cell(row=row, column=3, value=event['Location'])
            ws.cell(row=row, column=4, value=event['Entity_Type'])
            ws.cell(row=row, column=5, value=event['Entity_ID'])
            
            # Set row height to accommodate images
            ws.row_dimensions[row].height = 180
            
            # Save and insert initial frame
            img_buffer = io.BytesIO()
            event['Initial_Frame'].save(img_buffer, format='PNG')
            img = ExcelImage(img_buffer)
            ws.add_image(img, f'F{row}')
            
            # Save and insert final frame
            img_buffer = io.BytesIO()
            event['Final_Frame'].save(img_buffer, format='PNG')
            img = ExcelImage(img_buffer)
            ws.add_image(img, f'G{row}')
        
        # Create the output directory if it doesn't exist
        os.makedirs(output_path, exist_ok=True)
        
        # Save the workbook
        wb.save(full_path)
        print(f"Events with images exported to: {full_path}")
        return full_path

    def process_proximity_logic(self, detections, frame):#it is used for detecting if any trash is inrelation to an entity
        for track_id, track_data in self.tracking_data.items():
            if track_data['class_id'] not in [0, 2]:  # Process both humans and vehicles
                continue

            entity_center = track_data['coordinates'][-1]
            is_trash_near = False
            entity_id = track_id
            entity_type = track_data['type']

            for det in detections:
                if det['class_id'] == 1:  # Trash
                    trash_center = self.calculate_center(det['bbox'])
                    distance = self.compute_distance(entity_center, trash_center)
                    if distance < self.distance_threshold:
                        is_trash_near = True
                        break

            track_data['proximity_buffer'].append(1 if is_trash_near else 0)

            if self.frame_count % 4 == 0:
                track_data['frame_buffer'].append(frame.copy())

            consecutive_trash_frames = sum(track_data['proximity_buffer'])

            if track_data['state'] == 'IDLE':
                if consecutive_trash_frames >= self.min_holding_frames:
                    track_data['state'] = 'HOLDING_TRASH'
                    track_data['no_trash_count'] = 0
                    print(f"{entity_type.capitalize()} {entity_id} confirmed to be holding trash after {consecutive_trash_frames} frames")
            
            elif track_data['state'] == 'HOLDING_TRASH':
                if is_trash_near:
                    track_data['no_trash_count'] = 0
                else:
                    track_data['no_trash_count'] += 1
                    
                    if track_data['no_trash_count'] >= self.min_disposal_frames:
                        track_data['state'] = 'TRASH_DISPOSED'
                        print(f"{entity_type.capitalize()} {entity_id} disposed trash after {track_data['no_trash_count']} frames without trash")
                        self.save_evidence(entity_id, track_data['frame_buffer'], frame, entity_type)
                        track_data['frame_buffer'].clear()
                        track_data['proximity_buffer'] = deque([0] * self.temporal_window, maxlen=self.temporal_window)
                        track_data['no_trash_count'] = 0

    def clean_inactive_tracks(self, current_frame):
        inactive_tracks = [track_id for track_id, data in self.tracking_data.items()
                         if current_frame - data['last_seen'] > self.max_inactive_frames]
        for track_id in inactive_tracks:
            del self.tracking_data[track_id]
    
    def process_video(self, video_path, output_path):
        cap = cv2.VideoCapture(video_path)
        current_frame = 0

        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break
            
            new_width = int(frame.shape[1])
            new_height = int(frame.shape[0])
            frame = cv2.resize(frame, (new_width, new_height))
            
            detections = []
            
            # YOLOv8 for humans and vehicles
            yolo_results = self.yolo_model(frame, conf=0.5, iou=0.4)
            for result in yolo_results:
                for box in result.boxes:
                    bbox = box.xyxy[0].cpu().numpy()
                    class_id = int(box.cls[0].cpu().numpy())
                    confidence = float(box.conf[0].cpu().numpy())
                    
                    if class_id == 0:  # Person
                        detections.append({'bbox': bbox, 'class_id': 0})  # Human
                    elif class_id == 2:  # Car in COCO dataset
                        detections.append({'bbox': bbox, 'class_id': 2})  # Vehicle
                    elif class_id == 7:  # Truck in COCO dataset
                        detections.append({'bbox': bbox, 'class_id': 2})  # Vehicle
            
            # Custom model for trash detection
            trash_results = self.trash_model(frame, conf=0.5, iou=0.4)
            for result in trash_results:
                for box in result.boxes:
                    bbox = box.xyxy[0].cpu().numpy()
                    class_id = int(box.cls[0].cpu().numpy())
                    confidence = float(box.conf[0].cpu().numpy())
                    
                    if class_id == 1:  # Trash class from your model
                        detections.append({'bbox': bbox, 'class_id': 1})  # Trash

            self.assign_ids(detections, current_frame)
            self.process_proximity_logic(detections, frame)

            # Visualization
            for det in detections:
                x1, y1, x2, y2 = map(int, det['bbox'])
                if det['class_id'] == 1:  # Trash
                    color = (0, 255, 0)
                    label = f"ID {det.get('id', 'N/A')} - Trash"
                elif det['class_id'] == 0:  # Human
                    color = (0, 0, 255)
                    label = f"ID {det.get('id', 'N/A')} - Human"
                else:  # Vehicle
                    color = (255, 0, 0)
                    label = f"ID {det.get('id', 'N/A')} - Vehicle"
                
                cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

            y_offset = 30
            for track_id, track_data in self.tracking_data.items():
                if track_data['class_id'] in [0, 2]:  # Display states for both humans and vehicles
                    state = track_data['state']
                    entity_type = track_data['type'].capitalize()
                    display_text = f"{entity_type} ID {track_id}: {state}"
                    cv2.putText(frame, display_text, (10, y_offset), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
                    y_offset += 30

            cv2.imshow("Tracking", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

            current_frame += 1
            self.frame_count += 1

        cap.release()
        cv2.destroyAllWindows()
        
        # Export events to Excel at the end of video processing
        return self.export_events(output_path)


if __name__ == "__main__":
    # File paths
    trash_model_path = r"C:\Users\DELL\Desktop\trash\veh_pavi\vehidet\100epochv2.pt"
    video_path = r"C:\Users\DELL\Desktop\trash\veh_pavi\vehidet\mech1.mp4"
    evidence_path = r"C:\Users\DELL\Desktop\trash\veh_pavi\vehidet\evidence"
    excel_output_path = r"C:\Users\DELL\Desktop\trash\veh_pavi\vehidet\reports"
    
    # Initialize tracker with camera location
    tracker = WasteTracker(
        trash_model_path=trash_model_path,
        evidence_path=evidence_path,
        distance_threshold=150,
        pair_consistency_threshold=10,
        max_inactive_frames=90,
        temporal_window=10,
        temporal_threshold=7,
        min_holding_frames=5,
        min_disposal_frames=10,
        camera_location="Camera_01"  # Add camera location identifier
    )
    
    # Process video and export events
    excel_file = tracker.process_video(video_path, excel_output_path)
    
    if excel_file:
        print(f"\nProcessing complete!")
        print(f"Evidence saved in: {evidence_path}")
        print(f"Event log exported to: {excel_file}")
    else:
        print("\nProcessing complete! No disposal events detected.")